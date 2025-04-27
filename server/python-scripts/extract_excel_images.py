
import openpyxl
import os
import sys
import json
import base64
import re

def extract_images_from_excel(excel_file_path, output_dir):
    wb = openpyxl.load_workbook(excel_file_path)
    result = {"images": [], "error": None}
    
    try:
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        image_counter = 0
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Usar a forma oficial (embora possa ser lenta para alguns formatos) ou iterar por shapes
            # Vamos tentar iterar pelas imagens diretamente no sheet, que é mais comum
            if not hasattr(sheet, '_images'): continue
            
            print(f"Planilha '{sheet_name}' tem {len(sheet._images)} imagens.", file=sys.stderr)
            
            for image_obj in sheet._images:
                image_counter += 1
                image_data = None
                product_code = None

                # --- Tentativa de Acesso aos Dados da Imagem ---
                try:
                    # Em versões mais recentes, a imagem pode ser acessada diretamente
                    # O objeto openpyxl.drawing.image.Image tem um método _data() ou bytes diretamente?
                    # A documentação não é clara sobre o acesso direto aos bytes.
                    # O erro anterior sugere que _data pode ser um método.
                    
                    # Tentativa 1: Chamar _data() se for callable
                    if hasattr(image_obj, '_data') and callable(image_obj._data):
                        data_result = image_obj._data()
                        if isinstance(data_result, bytes):
                            image_data = data_result
                        else:
                             print(f"Aviso Img {image_counter}: image_obj._data() não retornou bytes.", file=sys.stderr)
                    
                    # Tentativa 2: Acessar _data diretamente se não for callable (e Tentativa 1 falhou)
                    elif hasattr(image_obj, '_data') and isinstance(image_obj._data, bytes):
                        image_data = image_obj._data

                    # Tentativa 3: Usar o método ref (pode ser caminho interno)
                    # Esta abordagem é mais complexa e menos garantida
                    # if not image_data and hasattr(image_obj, 'ref'):
                    #    print(f"Aviso Img {image_counter}: Tentando com image_obj.ref: {image_obj.ref}", file=sys.stderr)
                         # Precisaria abrir o arquivo zip do Excel e ler a referência interna

                    if not image_data:
                         print(f"Falha Img {image_counter}: Não foi possível obter dados binários válidos.", file=sys.stderr)
                         continue # Pula para a próxima imagem

                except Exception as data_err:
                    print(f"Erro Img {image_counter}: Erro ao acessar dados da imagem: {data_err}", file=sys.stderr)
                    continue # Pula para a próxima imagem
                # --- Fim da Tentativa de Acesso ---

                # Salvar temporariamente para converter (ainda necessário para base64)
                temp_image_name = f"temp_img_{image_counter}.png"
                temp_image_path = os.path.join(output_dir, temp_image_name)
                try:
                    with open(temp_image_path, "wb") as f:
                        # AQUI é onde o erro acontecia. image_data DEVE ser bytes.
                        if isinstance(image_data, bytes):
                            f.write(image_data)
                        else:
                             raise TypeError(f"image_data não é bytes, é {type(image_data)}")
                except Exception as write_err:
                     print(f"Erro Img {image_counter}: Erro ao salvar temp {temp_image_name}: {write_err}", file=sys.stderr)
                     continue 

                # Encontrar código do produto próximo (lógica existente mantida)
                try:
                    row = image_obj.anchor.to.row
                    col = image_obj.anchor.to.col
                    for r_offset in range(-3, 4):
                        for c_offset in range(-3, 4):
                             cell_row = max(1, row + r_offset)
                             cell_col = max(1, col + c_offset)
                             cell_value = sheet.cell(row=cell_row, column=cell_col).value
                             if cell_value and isinstance(cell_value, str):
                                 if cell_value.replace('.', '').isalnum() and len(cell_value) >= 5:
                                      product_code = cell_value
                                      break
                        if product_code: break
                except Exception as anchor_err:
                    print(f"Erro Img {image_counter}: Erro ao obter âncora/código: {anchor_err}", file=sys.stderr)
                
                if not product_code: product_code = f"unknown_product_{image_counter}"
                
                # Converter para base64
                encoded_image = None
                try:
                    with open(temp_image_path, "rb") as image_file:
                         encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
                except Exception as b64_err:
                    print(f"Erro Img {image_counter}: Erro ao converter para base64: {b64_err}", file=sys.stderr)
                finally:
                    # Garantir limpeza do arquivo temporário
                    if os.path.exists(temp_image_path):
                        try: os.remove(temp_image_path)
                        except: pass
                
                if not encoded_image:
                    continue # Pular se não conseguiu converter

                # Gerar nome final seguro
                safe_product_code = re.sub(r'[^w.-]', '_', str(product_code))
                image_filename = f"{safe_product_code}.png"
                
                # Adicionar ao resultado
                result["images"].append({
                    "product_code": product_code,
                    "image_filename": image_filename, 
                    "image_base64": encoded_image
                })
        
    except Exception as e:
        result["error"] = str(e)
        print(f"Erro geral na extração Python: {e}", file=sys.stderr)
    
    print(json.dumps(result))

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(json.dumps({"error": "Argumentos inválidos! Uso: python script.py arquivo_excel.xlsx diretório_saída"}))
        sys.exit(1)
    excel_file_path = sys.argv[1]
    output_dir = sys.argv[2]
    extract_images_from_excel(excel_file_path, output_dir)
