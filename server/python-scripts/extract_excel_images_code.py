
import openpyxl, os, sys, json, base64, re
# Importar tipos de âncora específicos
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, OneCellAnchor

def extract_images_from_excel(excel_file_path, output_dir):
    wb = openpyxl.load_workbook(excel_file_path)
    result = {"images": [], "error": None}
    try:
        if not os.path.exists(output_dir): os.makedirs(output_dir)
        image_counter = 0
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if not hasattr(sheet, '_images'): continue
            print(f"Planilha '{sheet_name}' tem {len(sheet._images)} imagens.", file=sys.stderr)
            for image_obj in sheet._images:
                image_counter += 1
                image_data = None
                product_code = None
                try:
                    # Tentar acessar dados da imagem
                    if hasattr(image_obj, '_data') and callable(image_obj._data):
                        data_result = image_obj._data()
                        if isinstance(data_result, bytes): image_data = data_result
                    elif hasattr(image_obj, '_data') and isinstance(image_obj._data, bytes):
                        image_data = image_obj._data
                    if not image_data: print(f"Falha Img {image_counter}: Dados binários inválidos.", file=sys.stderr); continue
                except Exception as data_err: print(f"Erro Img {image_counter} data: {data_err}", file=sys.stderr); continue

                temp_image_name = f"temp_img_{image_counter}.png"
                temp_image_path = os.path.join(output_dir, temp_image_name)
                try:
                    with open(temp_image_path, "wb") as f: f.write(image_data)
                except Exception as write_err: print(f"Erro Img {image_counter} save: {write_err}", file=sys.stderr); continue

                # --- LÓGICA AJUSTADA PARA OBTER ANCHOR ROW --- 
                anchor_row = -1 # Valor padrão se não encontrar
                try:
                    anchor = image_obj.anchor
                    # Verificar o TIPO da âncora
                    if isinstance(anchor, TwoCellAnchor):
                        # Se for TwoCellAnchor, usar .to.row
                        anchor_row = anchor.to.row + 1 
                        print(f"Img {image_counter}: Âncora TwoCellAnchor na linha {anchor_row}", file=sys.stderr)
                    elif isinstance(anchor, OneCellAnchor):
                        # Se for OneCellAnchor, usar .frm.row
                        anchor_row = anchor.frm.row + 1 
                        print(f"Img {image_counter}: Âncora OneCellAnchor na linha {anchor_row}", file=sys.stderr)
                    else:
                        print(f"Img {image_counter}: Tipo de âncora desconhecido: {type(anchor)}", file=sys.stderr)

                    # Tentar encontrar código na coluna F da linha da âncora
                    if anchor_row != -1:
                        code_cell_value = sheet.cell(row=anchor_row, column=6).value
                        if code_cell_value:
                            code_str = str(code_cell_value).strip()
                            if len(code_str) > 1 and not code_str.lower() in ['cod.', 'codigo', 'código']:
                                product_code = code_str
                                print(f"Img {image_counter}: Código da Col F linha {anchor_row}: '{product_code}'", file=sys.stderr)
                        if not product_code: # Fallback procurar perto
                             for offset in [-1, 1]:
                                 check_row = anchor_row + offset
                                 if check_row >= 1:
                                     fallback_cell = sheet.cell(row=check_row, column=6).value
                                     if fallback_cell:
                                         code_str = str(fallback_cell).strip()
                                         if len(code_str) > 1 and not code_str.lower() in ['cod.', 'codigo', 'código']:
                                             product_code = code_str
                                             print(f"Img {image_counter}: Código fallback Col F linha {check_row}: '{product_code}'", file=sys.stderr)
                                             break
                    if not product_code: product_code = f"unknown_product_{image_counter}"
                except Exception as anchor_err: print(f"Erro Img {image_counter} anchor: {anchor_err}", file=sys.stderr)
                # --- FIM DA LÓGICA DE ÂNCORA/CÓDIGO ---
                
                if not product_code: product_code = f"unknown_product_{image_counter}"
                print(f"Img {image_counter}: Código final: '{product_code}'", file=sys.stderr)

                encoded_image = None
                try:
                    with open(temp_image_path, "rb") as image_file: encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
                except Exception as b64_err: print(f"Erro Img {image_counter} base64: {b64_err}", file=sys.stderr)
                finally:
                    if os.path.exists(temp_image_path): 
                        try: os.remove(temp_image_path)
                        except OSError as e: print(f"Erro ao remover temp {temp_image_path}: {e}", file=sys.stderr)
                if not encoded_image: continue

                # Regex CORRIGIDA (hífen escapado no final)
                safe_product_code = re.sub(r'[^w.-]', '_', str(product_code))
                image_filename = f"{safe_product_code}.png"

                result["images"].append({
                    "product_code": product_code,
                    "image_filename": image_filename,
                    "image_base64": encoded_image
                })
    except Exception as e: result["error"] = str(e); print(f"Erro geral Python: {e}", file=sys.stderr)
    print(json.dumps(result))

if __name__ == "__main__":
    if len(sys.argv) != 3: 
        print(json.dumps({"error": "Argumentos inválidos!"}))
        sys.exit(1)
    extract_images_from_excel(sys.argv[1], sys.argv[2])
