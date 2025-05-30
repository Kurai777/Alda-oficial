#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import os
import sys
import json
import base64
import re  # Importação explícita para uso em todo o módulo
import traceback  # Para capturar stack traces
from PIL import Image
import io

def extract_products_fixed_columns(excel_file_path, output_dir):
    """
    Extrai produtos do Excel usando um mapeamento fixo de colunas:
    A (1): nome do produto (ex: "Sofá Home")
    B (2): local (ex: "2º Piso", "Depósito")
    C (3): fornecedor
    D (4): imagem (objeto gráfico)
    F (6): código do produto
    G (7): descrição
    L (12): valor total (preço do produto)
    
    Também extrai imagens associadas e retorna a estrutura de dados completa.
    """
    # Abrir o arquivo Excel
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    
    # Resultado para JSON
    result = {
        "products": [],
        "images": [],
        "errors": []
    }
    
    # Criar diretório de saída se não existir
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    try:
        # Para cada planilha no arquivo
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Extrair produtos da planilha
            try:
                row_products = extract_products_from_sheet(sheet)
                result["products"].extend(row_products)
                
                # Extrair imagens da planilha
                sheet_images = extract_images_from_sheet(sheet, output_dir)
                result["images"].extend(sheet_images)
                
                # Associar imagens com produtos por código
                associate_images_with_products(result["products"], result["images"])
                
            except Exception as e:
                result["errors"].append(f"Erro ao processar planilha {sheet_name}: {str(e)}")
    
    except Exception as e:
        result["errors"].append(f"Erro geral ao processar arquivo Excel: {str(e)}")
    
    # Retornar resultado como JSON
    print(json.dumps(result))
    return result

def extract_products_from_sheet(sheet):
    """
    Extrai produtos da planilha com mapeamento fixo de colunas:
    A (1): nome do produto (ex: "Sofá Home")
    B (2): local (ex: "2º Piso", "Depósito")
    C (3): fornecedor
    D (4): imagem (objeto gráfico)
    F (6): código do produto
    G (7): descrição
    L (12): valor total (preço do produto)
    """
    products = []
    print(f"Lendo produtos da planilha com {sheet.max_row} linhas...", file=sys.stderr)
    
    # Começar da linha 2 (pular cabeçalho)
    for row_idx in range(2, sheet.max_row + 1):
        try:
            # Ler valores das colunas específicas
            nome_cell = sheet.cell(row=row_idx, column=1)  # Coluna A (1): nome
            local_cell = sheet.cell(row=row_idx, column=2)  # Coluna B (2): local
            fornecedor_cell = sheet.cell(row=row_idx, column=3)  # Coluna C (3): fornecedor
            codigo_cell = sheet.cell(row=row_idx, column=6)  # Coluna F (6): código
            descricao_cell = sheet.cell(row=row_idx, column=7)  # Coluna G (7): descrição
            preco_cell = sheet.cell(row=row_idx, column=12)  # Coluna L (12): valor total
            
            # Obter valores das células
            nome = nome_cell.value if nome_cell else None
            local = local_cell.value if local_cell else None
            fornecedor = fornecedor_cell.value if fornecedor_cell else None
            codigo = codigo_cell.value if codigo_cell else None
            descricao = descricao_cell.value if descricao_cell else None
            preco = preco_cell.value if preco_cell else None
            
            # Limpar e normalizar valores (remover espaços extras, etc.)
            if nome and isinstance(nome, str):
                nome = nome.strip()
            if local and isinstance(local, str):
                local = local.strip()
            if fornecedor and isinstance(fornecedor, str):
                fornecedor = fornecedor.strip()
            if codigo and isinstance(codigo, str):
                codigo = codigo.strip()
            elif codigo:
                codigo = str(codigo)
            if descricao and isinstance(descricao, str):
                descricao = descricao.strip()
            
            # Debug: imprimir valores lidos
            if row_idx < 10 or row_idx % 50 == 0:  # Limitar log
                print(f"Linha {row_idx}: nome='{nome}', codigo='{codigo}', preco='{preco}'", file=sys.stderr)
            
            # Pular linhas vazias ou com nome/código vazios ou inválidos
            if (not nome or nome == "_EMPTY_" or nome == "" or 
                not codigo or codigo == "_EMPTY_" or codigo == "" or 
                codigo == "UNKNOWN-CODE"):
                if row_idx < 10 or row_idx % 50 == 0:  # Limitar log
                    print(f"  -> Pulando linha {row_idx} (nome ou código inválido)", file=sys.stderr)
                continue
            
            # Formatar preço como "R$ XX.XXX,XX"
            preco_formatado = format_price(preco)
            
            # Criar produto
            produto = {
                "nome": nome,
                "local": local or "",
                "fornecedor": fornecedor or "",
                "codigo": str(codigo).strip(),
                "descricao": descricao or "",
                "preco": preco_formatado,
                "imagem": "",  # Será preenchido depois ao associar imagens
                "row": row_idx  # Armazenar o número da linha para associar com imagens
            }
            
            products.append(produto)
        except Exception as e:
            print(f"Erro ao processar produto na linha {row_idx}: {str(e)}", file=sys.stderr)
    
    return products

def format_price(price_value):
    """
    Formata o valor do preço como "R$ XX.XXX,XX"
    """
    if not price_value:
        return "R$ 0,00"
    
    try:
        # Debug - imprimir o valor e tipo recebido
        print(f"Formatando preço: valor={price_value}, tipo={type(price_value)}", file=sys.stderr)
        
        # Converter para string se for outro tipo
        price_str = str(price_value)
        
        # Limpar a string de qualquer formatação existente
        # Remover R$, espaços e outros caracteres não essenciais
        cleaned_value = price_str.replace("R$", "").replace(" ", "").strip()
        
        # Verificar se é um número com formatação brasileira (1.234,56)
        if "," in cleaned_value and "." in cleaned_value:
            # Remover pontos de milhar e substituir vírgula por ponto para decimal
            cleaned_value = cleaned_value.replace(".", "").replace(",", ".")
        elif "," in cleaned_value:
            # Se só tem vírgula, substituir por ponto (formato decimal)
            cleaned_value = cleaned_value.replace(",", ".")
        
        # Tentar converter para float
        try:
            price_float = float(cleaned_value)
        except:
            # Se falhar, pode ser que esteja em notação científica ou outro formato
            import re
            # Extrair apenas dígitos e pontos/vírgulas
            nums = re.findall(r'[\d.,]+', cleaned_value)
            if nums:
                # Pegar o primeiro número encontrado
                num_str = nums[0].replace(",", ".")
                price_float = float(num_str)
            else:
                # Se não encontrou números, usar 0
                price_float = 0.0
        
        # Formatar como R$ XX.XXX,XX (formato brasileiro)
        formatted_price = f"R$ {price_float:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        
        print(f"Preço formatado: {formatted_price}", file=sys.stderr)
        return formatted_price
    
    except (ValueError, TypeError) as e:
        # Em caso de erro, registrar e retornar um valor padrão formatado
        print(f"Erro ao formatar preço '{price_value}': {str(e)}", file=sys.stderr)
        return f"R$ {price_value}"

def extract_images_from_sheet(sheet, output_dir):
    """
    Extrai imagens da planilha utilizando um método mais robusto.
    """
    images = []
    
    # Verificar se há imagens na planilha de diferentes formas
    try:
        # Método 1: Tentar acessar pelas imagens embutidas se disponível
        sheet_images = []
        
        # Tentar diferentes atributos para acessar imagens em diferentes versões do openpyxl
        image_attributes = ['_images', 'images', '_drawings', 'drawings']
        for attr in image_attributes:
            if hasattr(sheet, attr):
                try:
                    images_attr = getattr(sheet, attr)
                    if images_attr:
                        print(f"Encontrado atributo de imagens: {attr} com {len(images_attr)} imagens", file=sys.stderr)
                        sheet_images = images_attr
                        break
                except Exception as attr_err:
                    print(f"Erro ao acessar {attr}: {attr_err}", file=sys.stderr)
                    continue
        
        # Se não encontramos imagens, tentar métodos alternativos
        if not sheet_images:
            # Tentar acessar _drawing se disponível (em algumas versões do openpyxl)
            if hasattr(sheet, '_drawing') and sheet._drawing:
                try:
                    drawing_images = getattr(sheet._drawing, 'images', []) or []
                    if drawing_images:
                        print(f"Encontradas {len(drawing_images)} imagens via _drawing.images", file=sys.stderr)
                        sheet_images = drawing_images
                except Exception as drawing_err:
                    print(f"Erro ao acessar _drawing.images: {drawing_err}", file=sys.stderr)
        
        # Processar as imagens encontradas
        total_saved = 0
        print(f"Processando {len(sheet_images)} imagens encontradas na planilha", file=sys.stderr)
        
        for image_idx, image_tuple in enumerate(sheet_images):
            try:
                print(f"Processando imagem {image_idx}...", file=sys.stderr)
                
                # Gerar nome temporário para a imagem
                temp_image_name = f"temp_image_{image_idx}.png"
                temp_image_path = os.path.join(output_dir, temp_image_name)
                
                # Tentar diferentes formas de obter os dados da imagem
                image_data = None
                
                # Lista de atributos possíveis que podem conter dados binários da imagem
                binary_data_attrs = ['_data', 'data', 'blob', '_blob', 'content', '_content', 'image_data']
                
                # Tentar cada um dos atributos
                for attr in binary_data_attrs:
                    if hasattr(image_tuple, attr):
                        try:
                            data = getattr(image_tuple, attr)
                            if isinstance(data, bytes) and len(data) > 100:  # Pelo menos 100 bytes
                                image_data = data
                                print(f"Dados binários encontrados no atributo: {attr}", file=sys.stderr)
                                break
                        except Exception as bin_err:
                            print(f"Erro ao acessar {attr}: {bin_err}", file=sys.stderr)
                            continue
                
                # Se não encontrou pelos atributos diretos, tentar por ref/blob
                if not image_data and hasattr(image_tuple, 'ref') and hasattr(image_tuple.ref, 'blob'):
                    try:
                        blob_data = image_tuple.ref.blob
                        if isinstance(blob_data, bytes) and len(blob_data) > 100:
                            image_data = blob_data
                            print(f"Dados binários encontrados via ref.blob", file=sys.stderr)
                    except Exception as ref_err:
                        print(f"Erro ao acessar ref.blob: {ref_err}", file=sys.stderr)
                
                # Verificar se temos dados válidos
                if not image_data or not isinstance(image_data, bytes) or len(image_data) < 100:
                    print(f"Dados da imagem {image_idx} inválidos ou muito pequenos", file=sys.stderr)
                    # Continuar para a próxima imagem
                    continue
                
                # Salvar imagem em disco temporariamente
                with open(temp_image_path, "wb") as f:
                    f.write(image_data)
                
                # Obter informação da linha onde a imagem está
                row = 0
                col = 0
                
                # Tentar diferentes formas de obter as coordenadas da imagem
                try:
                    # Opção 1: pelo objeto anchor (tipo comum)
                    if hasattr(image_tuple, 'anchor'):
                        anchor = image_tuple.anchor
                        if hasattr(anchor, 'to'):
                            if hasattr(anchor.to, 'row'):
                                row = anchor.to.row
                            if hasattr(anchor.to, 'col'):
                                col = anchor.to.col
                        elif hasattr(anchor, 'row'):
                            row = anchor.row
                        elif hasattr(anchor, 'col'):
                            col = anchor.col
                    # Opção 2: pelas próprias propriedades
                    elif hasattr(image_tuple, 'row'):
                        row = image_tuple.row
                    elif hasattr(image_tuple, 'col'):
                        col = image_tuple.col
                except Exception as coord_err:
                    print(f"Erro ao obter coordenadas: {coord_err}", file=sys.stderr)
                
                print(f"Coordenadas da imagem {image_idx}: linha={row}, coluna={col}", file=sys.stderr)
                
                # Determinar o código do produto para esta imagem
                codigo = None
                
                # Estratégia 1: Se temos a linha, usar prioritariamente o código na coluna F (6)
                if row > 0:
                    # Verificar na mesma linha
                    codigo_cell = sheet.cell(row=row, column=6).value
                    if codigo_cell and str(codigo_cell).strip():
                        codigo = str(codigo_cell).strip()
                        print(f"Código encontrado na mesma linha {row}: {codigo}", file=sys.stderr)
                
                # Estratégia 2: Procurar pelas linhas próximas (3 linhas acima e abaixo)
                if not codigo and row > 0:
                    for r in range(max(1, row-3), min(sheet.max_row, row+3)):
                        codigo_cell = sheet.cell(row=r, column=6).value
                        if codigo_cell and str(codigo_cell).strip():
                            codigo = str(codigo_cell).strip()
                            print(f"Código encontrado em linha próxima {r}: {codigo}", file=sys.stderr)
                            break
                
                # Estratégia 3: Se a imagem estiver na coluna D (4), buscar código na mesma linha
                if not codigo and col == 4:
                    # Verificar se esta linha tem um código válido
                    codigo_cell = sheet.cell(row=row, column=6).value
                    if codigo_cell and str(codigo_cell).strip():
                        codigo = str(codigo_cell).strip()
                        print(f"Código encontrado para imagem na coluna D: {codigo}", file=sys.stderr)
                
                # Estratégia 4: Buscar todas as ocorrências de imagens na coluna D (4)
                if not codigo:
                    for r in range(2, sheet.max_row + 1):  # Começar da linha 2
                        # Se a célula na coluna D (4) não está vazia (possível local de imagem)
                        try:
                            cell_d = sheet.cell(row=r, column=4)
                            if cell_d and (cell_d.value or abs(r - row) < 2):  # mesmo que sem valor, se próxima da nossa linha
                                # Verificar se esta linha tem um código válido
                                codigo_cell = sheet.cell(row=r, column=6).value
                                if codigo_cell and str(codigo_cell).strip():
                                    codigo = str(codigo_cell).strip()
                                    print(f"Código encontrado na linha {r} com possível imagem: {codigo}", file=sys.stderr)
                                    row = r  # Atualizar a linha para esta
                                    break
                        except Exception as cell_err:
                            print(f"Erro ao verificar célula: {cell_err}", file=sys.stderr)
                            continue
                
                # Estratégia 5 (último recurso): Usar índice como identificador único
                if not codigo:
                    # Criar código baseado em índice, linha, coluna e timestamp para garantir exclusividade
                    import time
                    timestamp = int(time.time())
                    codigo = f"imagem_idx{image_idx}_r{row}_c{col}_{timestamp}"
                    print(f"Sem código identificado, usando identificador único: {codigo}", file=sys.stderr)
                
                # Remover caracteres inválidos do código para uso como nome de arquivo
                safe_codigo = re.sub(r'[^\w\-\.]', '_', codigo)
                
                # Adicionar o índice da imagem ao código para garantir exclusividade
                safe_codigo = f"{safe_codigo}_{image_idx}"
                
                # Definir nome e caminho final da imagem
                image_filename = f"{safe_codigo}.png"
                image_path = os.path.join(output_dir, image_filename)
                
                # Se já existe arquivo com esse nome, adicionar sufixo
                suffix = 1
                while os.path.exists(image_path):
                    image_filename = f"{safe_codigo}_{suffix}.png"
                    image_path = os.path.join(output_dir, image_filename)
                    suffix += 1
                
                # Verificar se a imagem é válida e tem um tamanho razoável
                try:
                    from PIL import Image
                    img = Image.open(temp_image_path)
                    width, height = img.size
                    
                    # Verificar dimensões - ignorar imagens muito pequenas (possíveis ícones/lixo)
                    if width < 20 or height < 20:
                        print(f"Imagem muito pequena ({width}x{height}), pulando...", file=sys.stderr)
                        img.close()
                        if os.path.exists(temp_image_path):
                            os.remove(temp_image_path)
                        continue
                    
                    # Fechar a imagem após verificação
                    img.close()
                except Exception as img_err:
                    print(f"Erro ao verificar imagem: {img_err}", file=sys.stderr)
                    # Continuar mesmo com erro, a imagem ainda pode ser válida
                
                # Copiar para o caminho final
                import shutil
                if os.path.exists(temp_image_path):
                    shutil.copy2(temp_image_path, image_path)
                    os.remove(temp_image_path)  # Remover temporário
                
                # Converter imagem para base64
                with open(image_path, "rb") as image_file:
                    encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
                
                # Adicionar à lista de imagens
                image_info = {
                    "codigo": codigo,
                    "filename": image_filename,
                    "path": image_path,
                    "base64": encoded_image,
                    "row": row,  # Guardar a linha para associação
                    "col": col   # Guardar a coluna também
                }
                
                images.append(image_info)
                total_saved += 1
                print(f"Imagem {image_idx} processada com sucesso: {image_filename}", file=sys.stderr)
                
            except Exception as e:
                print(f"Erro no script Python: Erro ao processar imagem {image_idx}: {str(e)}", file=sys.stderr)
                print(f"Stack trace: {traceback.format_exc()}", file=sys.stderr)
        
        print(f"Total de imagens extraídas e salvas: {total_saved} de {len(sheet_images)}", file=sys.stderr)
        
    except Exception as e:
        print(f"Erro geral no script Python: {str(e)}", file=sys.stderr)
        print(f"Stack trace: {traceback.format_exc()}", file=sys.stderr)
    
    # Segundo método: usar PIL para procurar imagens (fallback)
    if len(images) == 0:
        try:
            from PIL import Image as PILImage
            import tempfile
            
            # Salvar o arquivo Excel como temporário com extensão .xlsx
            temp_dir = tempfile.mkdtemp()
            temp_excel = os.path.join(temp_dir, "temp.xlsx")
            wb = openpyxl.load_workbook(filename=excel_file_path)
            wb.save(temp_excel)
            
            # Usar a biblioteca zipfile para extrair imagens
            import zipfile
            
            with zipfile.ZipFile(temp_excel, 'r') as zip_ref:
                # Extrair todas as imagens do arquivo
                image_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/')]
                
                for idx, image_file in enumerate(image_files):
                    try:
                        # Extrair a imagem para um arquivo temporário
                        img_temp = os.path.join(temp_dir, f"temp_img_{idx}.png")
                        with open(img_temp, 'wb') as f:
                            f.write(zip_ref.read(image_file))
                        
                        # Gerar nome único baseado no índice, timestamp e nome do arquivo original
                        import time
                        timestamp = int(time.time())
                        original_name = os.path.basename(image_file)
                        codigo = f"img_{idx}_{timestamp}_{original_name.replace('.', '_')}"
                        
                        # Garantir que o código seja único
                        safe_codigo = re.sub(r'[^\w\-\.]', '_', codigo)
                        image_filename = f"{safe_codigo}.png"
                        image_path = os.path.join(output_dir, image_filename)
                        
                        # Copiar para o destino final
                        import shutil
                        shutil.copy2(img_temp, image_path)
                        
                        # Converter para base64
                        with open(image_path, "rb") as image_file:
                            encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
                        
                        # Adicionar à lista
                        images.append({
                            "codigo": codigo,
                            "filename": image_filename,
                            "path": image_path,
                            "base64": encoded_image
                        })
                        
                    except Exception as e:
                        print(f"Erro no script Python: Erro ao processar imagem zip {idx}: {str(e)}", file=sys.stderr)
            
            # Limpar arquivos temporários
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)
            
        except Exception as e:
            print(f"Erro no script Python: Erro no método fallback: {str(e)}", file=sys.stderr)
    
    print(f"Total de {len(images)} imagens extraídas", file=sys.stderr)
    return images

def associate_images_with_products(products, images):
    """
    Associa imagens aos produtos com base no código e na linha, garantindo uma associação 1:1 estrita.
    """
    print(f"Associando {len(images)} imagens a {len(products)} produtos...", file=sys.stderr)
    
    # Criar dicionário de imagens por código
    images_by_code = {}
    for img in images:
        images_by_code[img["codigo"]] = img
    
    # Criar dicionário de imagens por linha
    images_by_row = {}
    for img in images:
        if "row" in img and img["row"] > 0:
            # Usar linha+coluna como chave única para garantir associação precisa
            row_col_key = f"{img['row']}_{img.get('col', 0)}"
            images_by_row[row_col_key] = img
    
    # Manter registro de imagens já utilizadas para evitar duplicação
    used_images = set()
    
    # Número de produtos atualizados
    updated_products = 0
    
    # Primeira passagem: Associação exata por código
    for product in products:
        codigo = product["codigo"]
        # Associar pelo código exato apenas
        if codigo in images_by_code and images_by_code[codigo]["codigo"] not in used_images:
            product["imagem"] = f"data:image/png;base64,{images_by_code[codigo]['base64']}"
            used_images.add(images_by_code[codigo]["codigo"])
            updated_products += 1
            print(f"[EXATO] Associada imagem ao produto '{product['nome']}' pelo código: {codigo}", file=sys.stderr)
            # Armazenar ID da imagem no produto para referência
            product["image_id"] = images_by_code[codigo]["codigo"]
    
    # Segunda passagem: Associação por linha e coluna para produtos sem imagem
    for product in products:
        if "imagem" in product:
            continue  # Já tem imagem da primeira passagem
            
        if "row" in product and product["row"] > 0:
            # Gerar chave linha+coluna
            row_col_key = f"{product['row']}_{product.get('col', 0)}"
            
            if row_col_key in images_by_row and images_by_row[row_col_key]["codigo"] not in used_images:
                product["imagem"] = f"data:image/png;base64,{images_by_row[row_col_key]['base64']}"
                used_images.add(images_by_row[row_col_key]["codigo"])
                updated_products += 1
                print(f"[LINHA] Associada imagem ao produto '{product['nome']}' pela linha/coluna: {row_col_key}", file=sys.stderr)
                # Armazenar ID da imagem
                product["image_id"] = images_by_row[row_col_key]["codigo"]
    
    # Terceira passagem: Associação por proximidade para produtos sem imagem
    # Usar um critério mais estrito para evitar falsas associações
    for product in products:
        if "imagem" in product:
            continue  # Já tem imagem
            
        codigo = product["codigo"]
        closest_img = None
        closest_img_codigo = None
        
        for img_code, img in images_by_code.items():
            if img_code in used_images:
                continue  # Imagem já usada
                
            # Associar apenas se o código do produto estiver contido na imagem ou vice-versa
            # E aplicar comparação de proximidade para escolher a melhor
            if codigo in img_code or img_code in codigo:
                if closest_img is None:
                    closest_img = img
                    closest_img_codigo = img_code
        
        if closest_img is not None:
            product["imagem"] = f"data:image/png;base64,{closest_img['base64']}"
            used_images.add(closest_img_codigo)
            updated_products += 1
            print(f"[PARCIAL] Associada imagem ao produto '{product['nome']}': {codigo} ↔ {closest_img_codigo}", file=sys.stderr)
            # Armazenar ID da imagem
            product["image_id"] = closest_img_codigo
    
    print(f"Associadas imagens a {updated_products} de {len(products)} produtos ({updated_products/len(products)*100:.1f}%)", file=sys.stderr)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(json.dumps({"error": "Argumentos inválidos! Uso: python script.py arquivo_excel.xlsx diretório_saída"}))
        sys.exit(1)
    
    excel_file_path = sys.argv[1]
    output_dir = sys.argv[2]
    
    extract_products_fixed_columns(excel_file_path, output_dir)