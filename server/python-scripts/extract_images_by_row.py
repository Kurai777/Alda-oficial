import openpyxl, os, sys, json, base64
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, OneCellAnchor

def extract_images_by_row(excel_file_path):
    """Extrai todas as imagens, a linha de suas âncoras E O NOME DA PLANILHA de todas as planilhas."""
    print(f"[Python DEBUG] Script recebendo caminho: {excel_file_path}", file=sys.stderr)
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        print(f"[Python DEBUG] Workbook carregado. Planilhas encontradas: {len(wb.sheetnames)} -> {wb.sheetnames}", file=sys.stderr)
    except Exception as load_err:
        print(f"[Python FATAL] Erro ao carregar workbook de {excel_file_path}: {load_err}", file=sys.stderr)
        print(json.dumps({"images": [], "error": f"Erro ao carregar workbook: {load_err}"}))
        sys.exit(1) 
        
    result = {"images": [], "error": None}
    images_processed_total = 0
    
    try:
        for sheet_name in wb.sheetnames:
            print(f"\n[Python DEBUG] Processando planilha: '{sheet_name}'", file=sys.stderr)
            sheet = wb[sheet_name]
            image_counter_sheet = 0
            images_processed_sheet = 0

            # DEBUG: SEMPRE verificar e logar _drawing count
            drawing_count = 0
            if hasattr(sheet, '_drawing') and sheet._drawing and hasattr(sheet._drawing, 'drawing_list'):
                drawing_count = len(sheet._drawing.drawing_list)
                if drawing_count > 0:
                    print(f"[Python DEBUG] Planilha '{sheet_name}' contém {drawing_count} objetos de desenho.", file=sys.stderr)
                # else: # Logar se não houver desenhos é muito verboso
                #    print(f"[Python DEBUG] Planilha '{sheet_name}' não contém objetos de desenho.", file=sys.stderr) 
            else:
                 print(f"[Python DEBUG] Atributo '_drawing' não encontrado ou vazio na planilha '{sheet_name}'.", file=sys.stderr)
            
            # Tentar processar _images se existir e não estiver vazio
            if hasattr(sheet, '_images') and len(sheet._images) > 0: 
                print(f"[Python DEBUG] Atributo '_images' existe. Verificando {len(sheet._images)} imagens em '{sheet_name}'...", file=sys.stderr)
                for image_obj in sheet._images:
                    image_counter_sheet += 1
                    image_data = None
                    anchor_row = -1
                    try:
                        if hasattr(image_obj, 'anchor') and hasattr(image_obj.anchor, '_from') and hasattr(image_obj.anchor._from, 'row'):
                           anchor_row = image_obj.anchor._from.row + 1 
                        else:
                           print(f"Warn: Img {image_counter_sheet} em '{sheet_name}' (tipo: {type(image_obj)}) sem âncora válida.", file=sys.stderr)
                           continue

                        if hasattr(image_obj, '_data') and callable(image_obj._data):
                            data = image_obj._data()
                            if isinstance(data, bytes):
                                image_data = data
                        elif hasattr(image_obj, '_data') and isinstance(image_obj._data, bytes):
                             image_data = image_obj._data

                        if image_data is None:
                            print(f"Warn: Img {image_counter_sheet} em '{sheet_name}' (Linha Âncora {anchor_row}) não foi possível extrair dados binários.", file=sys.stderr)
                            continue
                        
                        if image_data is not None and anchor_row != -1:
                            try:
                                encoded_image = base64.b64encode(image_data).decode('utf-8')
                                result["images"].append({
                                    "anchor_row": anchor_row,
                                    "image_base64": encoded_image,
                                    "sheet_name": sheet_name
                                })
                                images_processed_sheet += 1
                            except Exception as encode_err:
                                print(f"[Python WARN] Erro ao codificar Img {image_counter_sheet} da linha {anchor_row} (Planilha: {sheet_name}): {encode_err}", file=sys.stderr)

                    except Exception as img_err:
                         print(f"Erro ao processar Img {image_counter_sheet} em '{sheet_name}': {img_err}", file=sys.stderr)
            
            # Se _images estava vazio, mas _drawing tinha algo, logar isso (ainda sem extrair de _drawing)
            elif drawing_count > 0: 
                 print(f"[Python INFO] '_images' estava vazio, mas {drawing_count} objetos de desenho foram encontrados em '{sheet_name}' (extração de desenhos não implementada).", file=sys.stderr)
            else: # Nem _images nem _drawing
                 print(f"[Python WARN] Nenhuma imagem encontrada em '{sheet_name}' via '_images' ou '_drawing'.", file=sys.stderr)

            images_processed_total += images_processed_sheet
            if images_processed_sheet > 0:
                 print(f"[Python DEBUG] Planilha '{sheet_name}' concluída. {images_processed_sheet} imagens processadas a partir de '_images'.", file=sys.stderr)
            # else: # Não logar se não processou nada
            #    print(f"[Python DEBUG] Nenhuma imagem processada a partir de '_images' na planilha '{sheet_name}'.", file=sys.stderr)

        print(f"\nExtração Python FINALIZADA. Total de {images_processed_total} imagens processadas em todas as planilhas.", file=sys.stderr)

    except Exception as e:
        print(f"Erro GERAL na extração Python: {e}", file=sys.stderr)
        result["error"] = str(e)

    print(json.dumps(result))

if __name__ == "__main__":
    if len(sys.argv) > 1:
        excel_file_path = sys.argv[1]
        extract_images_by_row(excel_file_path)
    else:
        print(json.dumps({"images": [], "error": "Caminho do arquivo Excel não fornecido."})) 